import re
import time
from concurrent.futures import ThreadPoolExecutor
import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill

pattern = 'https://pan.baidu.com/s/\S+'  # 正则表达式
avail_fill = PatternFill(fill_type='solid', fgColor='CCEECC')  # 可用链接填充浅绿色
fail_fill = PatternFill(fill_type='solid', fgColor='AAAAAA')  # 失效链接填充灰色


# 检查URL有效性
def check(url) -> bool:
    html = requests.get(url).text
    soup = BeautifulSoup(html, 'html.parser')
    # 查找是否有share-error标签，有则无效
    # print("查找是否有share-error标签，有则无效")
    return soup.find('div', class_='share-error-left') is None


# 对一个行，检查该行，对于存在度盘链接的单元格，按检查结果填充颜色
# 该函数作为线程任务
def checkRow(sheet, row):
    for row in sheet.iter_rows(min_row=row, max_row=row):
        for cell in row:
            result = re.search(pattern, str(cell.value))
            if result is not None:
                print(result.group())
                if check(result.group()):
                    cell.fill = avail_fill
                else:
                    cell.fill = fail_fill


if __name__ == '__main__':
    # 不加 .xlsx 的文件路径
    # 实际运行只需要将这里修改为待检查的文件
    file = 'D:GameTables/手机游戏合集'
    # 加载文件和表格
    wb = openpyxl.load_workbook(file + '.xlsx')
    sheet = wb.worksheets[0]
    # 创建一个包含10条线程的线程池
    pool = ThreadPoolExecutor(max_workers=10)
    # 逐行扫描并向线程池提交该行的任务
    for row in range(1, sheet.max_row + 1):
        pool.submit(checkRow, sheet, row)
        print('已提交第' + str(row) + '行')
    pool.shutdown()
    wb.save(file + '(' + time.strftime("%Y%m%d", time.localtime()) + ').xlsx')
